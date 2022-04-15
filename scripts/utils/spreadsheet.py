import openpyxl


def _get_nft_spreadsheet_data(_path, _token_id):

    """
    Simple function how to transfer data from spreadsheet to Python.

    1) Read the spreadsheet.
    2) Loop through the each column in the first row. The values from that are
        the keys in data dictionary.
    3) At the same time loop through each column in the row specified by
        the token_id parameter (token_id == ID in spreadsheet).
    4) Assign key to value.
    5) Return the data dictionary.
    """

    workbook = openpyxl.load_workbook(_path)
    sheet = workbook.active

    # total_rows = sheet.max_row
    total_columns = sheet.max_column

    data = {}

    for i in range(1, total_columns + 1):
        k = sheet.cell(row=1, column=i)
        v = sheet.cell(row=_token_id + 1, column=i)

        if (
            k.value == "NAME"
            or k.value == "DESCRIPTION"
            or k.value == "CREATOR"
            or k.value == "ARTIST"
        ):
            data[k.value] = v.value
        else:
            data[k.value] = str(v.value).split(", ")

    return data
