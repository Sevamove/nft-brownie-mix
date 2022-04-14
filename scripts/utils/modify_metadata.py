from scripts.utils.pinata import upload_file
from scripts.utils.helper import dump_to_json, load_from_json
from scripts.utils.config import PATH, PINATA, HASHLIPS
from scripts.collectible.config import (
    ALTERNATIVE_IMAGE,
    SPREADSHEET,
    ALTERNATIVE_DATA,
    NFT_EXTERNAL_LINK,
    ADDITIONAL_METADATA,
)
import openpyxl


def main():
    modify_metadata()


def modify_metadata(_token_id: int = None):

    """
    Modify metadata that is generated by hashlips engine.
    Insert additional data from spreadsheet.
    Upload image and metadata to IPFS/Pinata if enabled.
    Return token_uri.
    """

    token_id = _token_id

    image_path = (
        PATH["images"] + f"/{token_id}.png"
        if HASHLIPS["enabled"]
        else PATH["images"] + f"/{ALTERNATIVE_IMAGE['file_name']}"
    )
    metadata_path = PATH["token_metadata"] + f"/{token_id}.json"
    token_uri_path = PATH["token_URIs"] + f"/{token_id}.json"

    metadata = load_from_json(metadata_path)
    token_uri = load_from_json(token_uri_path)

    print(f"Starting with tokenId: {token_id} ...")

    # Delete unnecessary keys made by hashlips engine.
    try:
        del metadata["dna"]
        del metadata["date"]
        del metadata["edition"]
        del metadata["compiler"]
    except KeyError:
        print(f"---KeyError occured. Working further on tokenId {token_id}---")

    # Inserting spreadsheet data to the metadata.
    if SPREADSHEET["enabled"]:
        ss_data = _get_nft_spreadsheet_data(PATH["spreadsheet"], token_id)

        metadata["name"] = ss_data["Name"]
        metadata["description"] = ss_data["Description"]

        if not SPREADSHEET["include_hashlips_generated_metadata_attributes"]:
            metadata["attributes"] = []

        for key, value in ss_data.items():
            if key in SPREADSHEET["trait_types"]:
                for v in value:  # loop through value list
                    metadata["attributes"].append(
                        {"trait_type": key, "value": v.capitalize()}
                    )
    else:
        metadata["name"] = ALTERNATIVE_DATA["name"] + f" #{token_id}"
        metadata["description"] = ALTERNATIVE_DATA["description"]

    # Inserting external link to the metadata.
    if NFT_EXTERNAL_LINK["enabled"]:
        metadata["external_link"] = _get_nft_external_link(token_id)

    # Inserting additional key/value to the metadata.
    if ADDITIONAL_METADATA:
        for k, v in ADDITIONAL_METADATA.items():
            metadata[k] = v

    if not PINATA["enabled"]:

        metadata["image"] = f"ipfs://YourImageUri/{token_id}.png"
        dump_to_json(metadata, metadata_path)

        token_uri[str(token_id)] = f"ipfs://YourTokenUri/{token_id}.json"
        dump_to_json(token_uri, token_uri_path)

    if PINATA["enabled"]:

        # metadata["image"] = upload_to_ipfs(image_path)
        metadata["image"] = upload_file(image_path)
        dump_to_json(metadata, metadata_path)

        # token_uri[str(token_id)] = upload_to_ipfs(metadata_path)
        token_uri[str(token_id)] = upload_file(metadata_path)
        dump_to_json(token_uri, token_uri_path)

    print(f"Finished with tokenId: {token_id}")

    return token_uri[str(token_id)]


def _get_nft_external_link(_token_id):
    if NFT_EXTERNAL_LINK["token_id"]:
        return NFT_EXTERNAL_LINK["url"] + str(_token_id)
    return NFT_EXTERNAL_LINK["url"]


def _get_nft_spreadsheet_data(_path, _token_id):

    """
    Simple function how to transfer data from spreadsheet to Python.

    1) Read the spreadsheet.
    2) Loop through the each column in the first row. The values from that are
        the keys in data dictionary.
    3) At the same time loop through each column in the row specified by
        the token_id parameter (token_id == ID in spreadsheet).
    4) Assign key to value.
    5) Return the data dictionary.
    """

    workbook = openpyxl.load_workbook(_path)
    sheet = workbook.active

    # total_rows = sheet.max_row
    total_columns = sheet.max_column

    data = {}

    for i in range(1, total_columns + 1):
        k = sheet.cell(row=1, column=i)
        v = sheet.cell(row=_token_id + 1, column=i)

        if k.value == "Name" or k.value == "Description":
            data[k.value] = v.value
        else:
            data[k.value] = str(v.value).split(", ")

    return data