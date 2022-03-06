import openpyxl

def main():
    _get_nft_spreadsheet_data("./test.xlsx", 1)

def _get_nft_spreadsheet_data(_path, _id):
    workbook = openpyxl.load_workbook(_path)
    sheet = workbook.active

    total_rows = sheet.max_row
    total_columns = sheet.max_column
    #print("Total Rows:", total_rows)
    #print("Total Columns:", total_columns)

    #data = {
    #    "id": "",
    #    "name": "",
    #    "description": "",
    #    "attributes": [
    #        {"trait_type": "sport", "value": ""},
    #        {"trait_type": "hobby", "value": ""},
    #        {"trait_type": "speaks", "value": ""},
    #        {"trait_type": "character", "value": ""},
    #        {"trait_type": "location", "value": ""},
    #    ]
    #}

    data = []

    print("\nValue of first column")
    for i in range(1, total_rows + 1):
        cell_obj = sheet.cell(row=i, column=1)
        print(cell_obj.value)

    print("\nValue of first row")
    for i in range(1, total_columns + 1):
        cell = sheet.cell(row=_id+1, column=i)
        data.append(cell.value)
    print(data)

